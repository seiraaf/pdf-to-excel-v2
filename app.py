from io import BytesIO
from datetime import datetime
import re
import textwrap
from urllib.parse import quote

import pandas as pd
import pdfplumber
import streamlit as st
from PIL import Image, ImageDraw, ImageFont
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SUPPLIER_DB_PATH = r"C:\Users\TRM\Downloads\DATABASE SUP.xlsx"
UNITS = ["Nirwana", "Lovina", "Lembongan", "The Club", "Kanaka"]
COLUMNS = ["Unit", "Supplier", "No PO", "Item", "Qty", "Unit Item", "Price", "Amount", "Remarks"]
EXPORT_HEADERS = ["Supplier", "No PO", "Inv Name", "Q", "Unit", "Description", "Qty Datang", "Remark"]
UNIT_TITLES = {
    "Nirwana": "NIRWANA BEACH RESORT",
    "Lovina": "LOVINA BEACH RESORT",
    "Lembongan": "LEMBONGAN",
    "The Club": "THE CLUB",
    "Kanaka": "KANAKA",
}
MONTHS_ID = {
    "jan": "JANUARI",
    "january": "JANUARI",
    "feb": "FEBRUARI",
    "february": "FEBRUARI",
    "mar": "MARET",
    "march": "MARET",
    "apr": "APRIL",
    "april": "APRIL",
    "may": "MEI",
    "jun": "JUNI",
    "june": "JUNI",
    "jul": "JULI",
    "july": "JULI",
    "aug": "AGUSTUS",
    "august": "AGUSTUS",
    "sep": "SEPTEMBER",
    "sept": "SEPTEMBER",
    "september": "SEPTEMBER",
    "oct": "OKTOBER",
    "october": "OKTOBER",
    "nov": "NOVEMBER",
    "november": "NOVEMBER",
    "dec": "DESEMBER",
    "december": "DESEMBER",
}

NUMBER_PATTERN = r"\d{1,3}(?:,\d{3})*(?:\.\d+)?|\d+(?:\.\d+)?"
PATTERN_WITH_NO = re.compile(
    rf"^\s*\d+\s+(?P<item>.*?)\s+(?P<qty>{NUMBER_PATTERN})\s+(?P<unit>[A-Za-z]+)\s+"
    rf"(?P<price>{NUMBER_PATTERN})\s+(?P<amount>{NUMBER_PATTERN})(?:\s+(?P<tail>.*))?\s*$",
    re.IGNORECASE,
)
PATTERN_WITHOUT_NO = re.compile(
    rf"^\s*(?P<item>.*?)\s+(?P<qty>{NUMBER_PATTERN})\s+(?P<unit>[A-Za-z]+)\s+"
    rf"(?P<price>{NUMBER_PATTERN})\s+(?P<amount>{NUMBER_PATTERN})(?:\s+(?P<tail>.*))?\s*$",
    re.IGNORECASE,
)

SKIP_KEYWORDS = (
    "subtotal",
    "grand total",
    "purchase order",
    "prepared by",
    "approved by",
    "discount",
    "tax & freight",
    "total :",
)
TABLE_SETTINGS = {
    "vertical_strategy": "text",
    "horizontal_strategy": "text",
    "snap_tolerance": 4,
    "join_tolerance": 4,
    "intersection_tolerance": 6,
    "text_tolerance": 3,
}


st.set_page_config(page_title="PDF to Excel", page_icon="chart", layout="wide")


if "all_data" not in st.session_state:
    st.session_state.all_data = []

if "excel_file" not in st.session_state:
    st.session_state.excel_file = None


def clean_text(value, default="-"):
    value = re.sub(r"\s+", " ", str(value or "")).strip()
    return value if value else default


def normalize_number(value):
    return clean_text(value).replace(" ", "")


def to_excel_number(value):
    value = clean_text(value, "").replace(",", "")
    try:
        number = float(value)
    except ValueError:
        return clean_text(value)
    return int(number) if number.is_integer() else number


def normalize_lookup(value):
    value = clean_text(value, "").upper()
    value = value.replace("&", " AND ")
    value = re.sub(r"[^A-Z0-9]+", " ", value)
    return clean_text(value, "")


def compact_lookup(value):
    return normalize_lookup(value).replace(" ", "")


@st.cache_data(show_spinner=False)
def prepare_supplier_database(file_bytes=None, path=None):
    """Read item→vendor database and optionally a WhatsApp number column.

    Required columns: NAMA, VENDOR
    Optional WA aliases: NO WA, NO_WA, WA, WHATSAPP, PHONE, TELEPON
    """
    base_columns = ["NAMA", "VENDOR", "NO WA", "LOOKUP", "LOOKUP_COMPACT"]
    try:
        if file_bytes:
            df = pd.read_excel(BytesIO(file_bytes))
        else:
            df = pd.read_excel(path)
    except Exception:
        return pd.DataFrame(columns=base_columns)

    df.columns = [clean_text(column, "").upper() for column in df.columns]
    if "NAMA" not in df.columns or "VENDOR" not in df.columns:
        return pd.DataFrame(columns=base_columns)

    phone_aliases = ["NO WA", "NO_WA", "WA", "WHATSAPP", "PHONE", "TELEPON", "NO TELP"]
    phone_column = next((column for column in phone_aliases if column in df.columns), None)
    keep_columns = ["NAMA", "VENDOR"] + ([phone_column] if phone_column else [])
    df = df[keep_columns].copy()
    if phone_column and phone_column != "NO WA":
        df = df.rename(columns={phone_column: "NO WA"})
    if "NO WA" not in df.columns:
        df["NO WA"] = ""

    df = df.dropna(subset=["NAMA", "VENDOR"])
    df["NAMA"] = df["NAMA"].map(lambda value: clean_text(value, ""))
    df["VENDOR"] = df["VENDOR"].map(lambda value: clean_text(value, ""))
    df["NO WA"] = df["NO WA"].map(lambda value: clean_text(value, ""))
    df["LOOKUP"] = df["NAMA"].map(normalize_lookup)
    df["LOOKUP_COMPACT"] = df["NAMA"].map(compact_lookup)
    df = df[df["LOOKUP"] != ""]
    return df


def lookup_supplier(item, supplier_db):
    item_lookup = normalize_lookup(item)
    item_compact = compact_lookup(item)
    if not item_lookup or supplier_db.empty:
        return ""

    exact = supplier_db[supplier_db["LOOKUP"] == item_lookup]
    if not exact.empty:
        return exact.iloc[0]["VENDOR"]

    compact_exact = supplier_db[supplier_db["LOOKUP_COMPACT"] == item_compact]
    if not compact_exact.empty:
        return compact_exact.iloc[0]["VENDOR"]

    candidates = []
    for row in supplier_db.itertuples(index=False):
        db_item = row.LOOKUP
        db_compact = row.LOOKUP_COMPACT
        if len(db_compact) < 4:
            continue
        if db_item in item_lookup or item_lookup in db_item:
            candidates.append((len(db_item), row.VENDOR))
        elif db_compact in item_compact or item_compact in db_compact:
            candidates.append((len(db_compact), row.VENDOR))

    if candidates:
        return sorted(candidates, reverse=True)[0][1]
    return ""


def normalize_whatsapp_number(value):
    """Normalize Indonesian/mobile phone values for wa.me links."""
    raw = clean_text(value, "")
    if not raw:
        return ""

    # Excel may read numeric phone cells as 62812....0
    raw = re.sub(r"\.0$", "", raw)
    digits = re.sub(r"\D", "", raw)
    if not digits:
        return ""
    if digits.startswith("00"):
        digits = digits[2:]
    if digits.startswith("0"):
        digits = "62" + digits[1:]
    elif digits.startswith("8"):
        digits = "62" + digits
    return digits


def lookup_supplier_phone(supplier, supplier_db):
    if supplier_db.empty or not supplier or "NO WA" not in supplier_db.columns:
        return ""
    target = normalize_lookup(supplier)
    matches = supplier_db[supplier_db["VENDOR"].map(normalize_lookup) == target]
    if matches.empty:
        return ""
    for value in matches["NO WA"].tolist():
        phone = normalize_whatsapp_number(value)
        if phone:
            return phone
    return ""


def build_whatsapp_message(unit, supplier, group_df):
    po_date = next((clean_text(v, "") for v in group_df.get("PO Date", []) if clean_text(v, "")), "")
    no_po_values = [clean_text(v, "") for v in group_df.get("No PO", []) if clean_text(v, "")]
    no_po = no_po_values[0] if no_po_values else ""

    lines = [f"Halo {supplier},", "", "Mohon diproses pesanan berikut:", f"Unit: {UNIT_TITLES.get(unit, unit.upper())}"]
    if po_date:
        lines.append(po_date)
    if no_po:
        lines.append(f"No PO: {no_po}")
    lines.append("")

    for idx, row in enumerate(group_df.to_dict("records"), 1):
        item = clean_text(row.get("Item", ""), "")
        qty = clean_text(row.get("Qty", ""), "")
        unit_item = clean_text(row.get("Unit Item", ""), "")
        lines.append(f"{idx}. {item} — {qty} {unit_item}")

    lines.extend(["", "Detail terlampir pada gambar PO. Terima kasih."])
    return "\n".join(lines)


def _load_font(size, bold=False):
    candidates = (
        ["/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", "/usr/share/fonts/truetype/liberation2/LiberationSans-Bold.ttf"]
        if bold
        else ["/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", "/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf"]
    )
    for path in candidates:
        try:
            return ImageFont.truetype(path, size=size)
        except OSError:
            pass
    return ImageFont.load_default()


def _wrap_cell(text, max_chars):
    text = clean_text(text, "")
    if not text:
        return [""]
    return textwrap.wrap(text, width=max_chars, break_long_words=False, break_on_hyphens=False) or [text]


def build_supplier_po_image(unit, supplier, group_df):
    """Create a shareable PNG resembling the purchasing sheet shown by the user."""
    group_df = group_df.reset_index(drop=True)
    po_date = next((clean_text(v, "") for v in group_df.get("PO Date", []) if clean_text(v, "")), "PO")

    width = 1500
    margin = 24
    title_h = 66
    date_h = 48
    header_h = 58
    padding_y = 14
    line_h = 30

    # Supplier, No PO, Inv Name, Q, Unit — matches the working Excel layout.
    col_widths = [260, 250, 650, 160, 130]
    headers = ["Supplier", "No PO", "Inv Name", "Q", "Unit"]
    char_limits = [22, 20, 48, 12, 10]

    rows = []
    row_heights = []
    for row in group_df.to_dict("records"):
        values = [
            supplier,
            clean_text(row.get("No PO", ""), ""),
            clean_text(row.get("Item", ""), ""),
            clean_text(row.get("Qty", ""), ""),
            clean_text(row.get("Unit Item", ""), ""),
        ]
        wrapped = [_wrap_cell(v, lim) for v, lim in zip(values, char_limits)]
        lines = max(len(part) for part in wrapped)
        rows.append(wrapped)
        row_heights.append(max(56, lines * line_h + padding_y * 2))

    height = margin * 2 + title_h + date_h + 18 + header_h + sum(row_heights)
    img = Image.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(img)
    title_font = _load_font(38, bold=True)
    date_font = _load_font(25, bold=True)
    header_font = _load_font(23, bold=True)
    body_font = _load_font(22, bold=False)

    # Title + PO date
    y = margin
    draw.text((margin, y + 4), UNIT_TITLES.get(unit, unit.upper()), font=title_font, fill="black")
    y += title_h
    draw.text((margin, y + 4), po_date or "PO", font=date_font, fill="black")
    y += date_h + 18

    table_x = margin
    table_w = sum(col_widths)
    # Header background
    draw.rectangle((table_x, y, table_x + table_w, y + header_h), fill="#F2F2F2", outline="black", width=2)
    x = table_x
    for header, cw in zip(headers, col_widths):
        draw.rectangle((x, y, x + cw, y + header_h), outline="black", width=2)
        bbox = draw.textbbox((0, 0), header, font=header_font)
        tw = bbox[2] - bbox[0]
        th = bbox[3] - bbox[1]
        draw.text((x + (cw - tw) / 2, y + (header_h - th) / 2 - 3), header, font=header_font, fill="black")
        x += cw
    y += header_h

    for wrapped, rh in zip(rows, row_heights):
        x = table_x
        for col_idx, (parts, cw) in enumerate(zip(wrapped, col_widths)):
            draw.rectangle((x, y, x + cw, y + rh), outline="black", width=2)
            total_text_h = len(parts) * line_h
            ty = y + max(padding_y, (rh - total_text_h) / 2)
            for part in parts:
                bbox = draw.textbbox((0, 0), part, font=body_font)
                tw = bbox[2] - bbox[0]
                if col_idx in (3, 4):
                    tx = x + (cw - tw) / 2
                else:
                    tx = x + 10
                draw.text((tx, ty), part, font=body_font, fill="black")
                ty += line_h
            x += cw
        y += rh

    output = BytesIO()
    img.save(output, format="PNG", optimize=True)
    output.seek(0)
    return output


def looks_like_number(value):
    return bool(re.fullmatch(NUMBER_PATTERN, clean_text(value, "")))


def looks_like_unit(value):
    value = clean_text(value, "")
    return bool(re.fullmatch(r"[A-Za-z]{1,8}", value))


def should_skip_line(line):
    lowered = line.lower()
    return not line or any(keyword in lowered for keyword in SKIP_KEYWORDS)


def format_po_date(raw_date):
    raw_date = clean_text(raw_date, "")
    match = re.search(r"(\d{1,2})[-/\s]([A-Za-z]+)[-/\s](\d{2,4})", raw_date)
    if not match:
        return ""

    day, month, year = match.groups()
    if len(year) == 2:
        year = "20" + year

    month_name = MONTHS_ID.get(month.lower(), month.upper())
    return f"PO {int(day)} {month_name} {year}"


def extract_po_date(text):
    for line in (text or "").splitlines():
        match = re.search(r"\b(?:Date|Tanggal)\b\s*:?[\s-]*(.+)$", line, re.IGNORECASE)
        if match:
            po_date = format_po_date(match.group(1))
            if po_date:
                return po_date
    return ""


def extract_no_po(text):
    match = re.search(r"\bPO/\d+/\d+\b", text or "", re.IGNORECASE)
    return clean_text(match.group(0), "") if match else ""


def extract_header_remark(text):
    for line in (text or "").splitlines():
        match = re.search(r"\bRemark\b\s*:?[\s-]*(.+)$", line, re.IGNORECASE)
        if match:
            remark = clean_text(match.group(1), "")
            if remark and not remark.lower().startswith(("price", "amount", "supplier")):
                return remark
    return "-"


def extract_row_remark(tail, header_remark="-"):
    tail = clean_text(tail, "")
    header_remark = clean_text(header_remark)
    if not tail:
        return header_remark

    # Format PO kedua: Amount Remark PR No.
    tail_without_pr = re.split(r"\bPR/\d+/\d+\b", tail, maxsplit=1, flags=re.IGNORECASE)[0]
    tail_without_pr = clean_text(tail_without_pr, "")
    if tail_without_pr:
        return tail_without_pr

    # Format Daily Market List biasanya supplier ada setelah Amount, remarks ada di header.
    return header_remark


def make_row(unit, item, qty, unit_item, price, amount, remarks, no_po, supplier_db):
    return {
        "Unit": unit,
        "Supplier": lookup_supplier(item, supplier_db),
        "No PO": no_po,
        "Item": clean_text(item),
        "Qty": normalize_number(qty),
        "Unit Item": clean_text(unit_item).upper(),
        "Price": normalize_number(price),
        "Amount": normalize_number(amount),
        "Remarks": clean_text(remarks),
    }


def parse_line(line, unit, header_remark="-", po_date="", no_po="", supplier_db=None):
    line = clean_text(line, "")
    if should_skip_line(line):
        return None

    match = PATTERN_WITH_NO.match(line) or PATTERN_WITHOUT_NO.match(line)
    if not match:
        return None

    data = match.groupdict()
    remarks = extract_row_remark(data.get("tail", ""), header_remark)
    row = make_row(
        unit,
        data["item"],
        data["qty"],
        data["unit"],
        data["price"],
        data["amount"],
        remarks,
        no_po,
        supplier_db,
    )
    row["PO Date"] = po_date
    return row


def parse_table_row(cells, unit, header_remark="-", po_date="", no_po="", supplier_db=None):
    cells = [clean_text(cell, "") for cell in cells]
    cells = [cell for cell in cells if cell]
    line = " ".join(cells)

    if should_skip_line(line) or "item name" in line.lower() or len(cells) < 4:
        return None

    number_indexes = [index for index, cell in enumerate(cells) if looks_like_number(cell)]
    if len(number_indexes) < 3:
        return None

    qty_index = number_indexes[0]
    amount_index = number_indexes[-1]
    price_index = number_indexes[-2]

    unit_index = None
    for index in range(qty_index + 1, price_index):
        if looks_like_unit(cells[index]):
            unit_index = index
            break

    if unit_index is None:
        return None

    item = clean_text(" ".join(cells[:qty_index]), "")
    if not item:
        return None

    tail = clean_text(" ".join(cells[amount_index + 1 :]), "")
    remarks = extract_row_remark(tail, header_remark)
    row = make_row(
        unit,
        item,
        cells[qty_index],
        cells[unit_index],
        cells[price_index],
        cells[amount_index],
        remarks,
        no_po,
        supplier_db,
    )
    row["PO Date"] = po_date
    return row


def row_key(row):
    return (
        row.get("Unit", ""),
        row.get("No PO", ""),
        compact_lookup(row.get("Item", "")),
        normalize_number(row.get("Qty", "")),
        clean_text(row.get("Unit Item", "")).upper(),
        normalize_number(row.get("Price", "")),
        normalize_number(row.get("Amount", "")),
    )


def item_spacing_score(item):
    item = clean_text(item, "")
    tokens = item.split()
    short_tokens = sum(1 for token in tokens if len(token) <= 2)
    return (short_tokens, len(tokens), len(item))


def prefer_new_row(existing, new_row):
    if not existing:
        return True
    return item_spacing_score(new_row.get("Item", "")) < item_spacing_score(existing.get("Item", ""))


def merge_row(existing, new_row):
    if prefer_new_row(existing, new_row):
        merged = dict(new_row)
        fallback = existing
    else:
        merged = dict(existing)
        fallback = new_row

    for field in ["Supplier", "No PO", "Remarks", "PO Date"]:
        if clean_text(merged.get(field, ""), "") in ["", "-", "Unknown"]:
            merged[field] = fallback.get(field, merged.get(field, ""))
    return merged


def add_unique_row(rows_by_key, order, row):
    key = row_key(row)
    if key not in rows_by_key:
        rows_by_key[key] = row
        order.append(key)
        return
    rows_by_key[key] = merge_row(rows_by_key[key], row)


def parse_pdf(uploaded_file, unit, supplier_db):
    rows_by_key = {}
    order = []

    with pdfplumber.open(uploaded_file) as pdf:
        for page in pdf.pages:
            text = page.extract_text(x_tolerance=2, y_tolerance=3) or ""
            header_remark = extract_header_remark(text)
            po_date = extract_po_date(text)
            no_po = extract_no_po(text)

            for table in page.extract_tables(TABLE_SETTINGS) or []:
                for cells in table:
                    row = parse_table_row(cells, unit, header_remark, po_date, no_po, supplier_db)
                    if row:
                        add_unique_row(rows_by_key, order, row)

            for line in text.splitlines():
                row = parse_line(line, unit, header_remark, po_date, no_po, supplier_db)
                if row:
                    add_unique_row(rows_by_key, order, row)

    return [rows_by_key[key] for key in order]


def export_row(row):
    return [
        row.get("Supplier", ""),
        row.get("No PO", ""),
        row.get("Item", ""),
        to_excel_number(row.get("Qty", "")),
        row.get("Unit Item", ""),
        "",
        "",
        "" if row.get("Remarks") == "-" else row.get("Remarks", ""),
    ]


def style_sheet(ws, unit_name, po_date, last_row):
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="F2F2F2")

    ws.merge_cells("A1:H1")
    ws.merge_cells("A2:H2")
    ws["A1"] = UNIT_TITLES.get(unit_name, unit_name.upper())
    ws["A2"] = po_date or "PO"
    ws["A1"].font = Font(bold=True, size=20)
    ws["A2"].font = Font(bold=True, size=10)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 8
    ws.row_dimensions[4].height = 18

    widths = [18, 18, 44, 10, 8, 48, 16, 20]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width

    for cell in ws[4]:
        cell.font = Font(bold=True, size=9)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=5, max_row=max(last_row, 5), min_col=1, max_col=8):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.font = Font(size=10)
        row[2].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        row[3].alignment = Alignment(horizontal="right", vertical="center")
        row[4].alignment = Alignment(horizontal="center", vertical="center")
        row[3].number_format = "#,##0.00"

    ws.auto_filter.ref = f"A4:H{max(last_row, 5)}"
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = True


def build_excel(data):
    output = BytesIO()
    df = pd.DataFrame(data)
    workbook = Workbook()
    workbook.remove(workbook.active)

    for unit_name in UNITS:
        ws = workbook.create_sheet(unit_name[:31])
        unit_df = df[df["Unit"] == unit_name] if not df.empty and "Unit" in df.columns else pd.DataFrame()
        po_dates = []
        if not unit_df.empty and "PO Date" in unit_df.columns:
            po_dates = [clean_text(value, "") for value in unit_df["PO Date"] if clean_text(value, "")]
        po_date = po_dates[0] if po_dates else ""

        ws.append([""] * len(EXPORT_HEADERS))
        ws.append([""] * len(EXPORT_HEADERS))
        ws.append([""] * len(EXPORT_HEADERS))
        ws.append(EXPORT_HEADERS)

        if not unit_df.empty:
            for row in unit_df.to_dict("records"):
                ws.append(export_row(row))

        style_sheet(ws, unit_name, po_date, ws.max_row)

    workbook.save(output)
    output.seek(0)
    return output


# ---------- UI / APP LAYER ----------
st.markdown(
    """
    <style>
        .block-container {padding-top: 2rem; padding-bottom: 3rem; max-width: 1180px;}
        [data-testid="stSidebar"] {background: #f7f8fb;}
        .hero {
            padding: 1.25rem 1.4rem;
            border: 1px solid #e8eaf0;
            border-radius: 18px;
            background: linear-gradient(135deg, #ffffff 0%, #f7f9ff 100%);
            margin-bottom: 1rem;
        }
        .hero h1 {margin: 0; font-size: 2.15rem; line-height: 1.15;}
        .hero p {margin: .45rem 0 0 0; color: #667085;}
        .soft-card {
            padding: .85rem 1rem;
            border: 1px solid #eaecf0;
            border-radius: 14px;
            background: white;
        }
        div[data-testid="stMetric"] {
            border: 1px solid #eaecf0;
            border-radius: 14px;
            padding: .8rem 1rem;
            background: #ffffff;
        }
        .tiny {color:#667085; font-size:.88rem;}
    </style>
    """,
    unsafe_allow_html=True,
)

st.markdown(
    """
    <div class="hero">
        <h1>📄 PDF → Excel Converter</h1>
        <p>Upload PO per unit, cek hasilnya, rapikan bila perlu, lalu export menjadi 1 file Excel multi-sheet.</p>
    </div>
    """,
    unsafe_allow_html=True,
)

with st.sidebar:
    st.header("⚙️ Database Supplier")
    supplier_file = st.file_uploader(
        "Upload DATABASE SUP.xlsx",
        type=["xlsx"],
        key="supplier_db",
        help="Wajib: NAMA dan VENDOR. Opsional: NO WA untuk tombol WhatsApp langsung.",
    )
    if supplier_file:
        supplier_db = prepare_supplier_database(supplier_file.getvalue())
    else:
        supplier_db = prepare_supplier_database(path=SUPPLIER_DB_PATH)

    if supplier_db.empty:
        st.warning("Database supplier belum terbaca. Supplier tetap bisa diproses, tetapi kolom supplier dapat kosong.")
    else:
        st.success(f"{len(supplier_db):,} item supplier terbaca.")
        if "NO WA" in supplier_db.columns and supplier_db["NO WA"].astype(str).str.strip().ne("").any():
            phone_count = supplier_db.loc[supplier_db["NO WA"].astype(str).str.strip().ne(""), "VENDOR"].nunique()
            st.caption(f"📱 Nomor WhatsApp tersedia untuk {phone_count} supplier.")
        else:
            st.caption("Tambahkan kolom **NO WA** agar tombol chat supplier aktif.")

    st.divider()
    st.subheader("📊 Status Data")
    if st.session_state.all_data:
        status_df = pd.DataFrame(st.session_state.all_data)
        counts = status_df.groupby("Unit").size().reindex(UNITS, fill_value=0)
    else:
        counts = pd.Series(0, index=UNITS)

    total_rows = int(counts.sum())
    for unit_name, total in counts.items():
        pct = int((total / total_rows) * 100) if total_rows else 0
        st.write(f"**{unit_name}** · {total} baris")
        st.progress(pct / 100 if total_rows else 0)

    st.divider()
    if st.session_state.all_data:
        clear_unit = st.selectbox("Hapus data unit", ["-"] + UNITS, key="clear_unit")
        if clear_unit != "-" and st.button("Hapus Unit Terpilih", use_container_width=True):
            st.session_state.all_data = [r for r in st.session_state.all_data if r.get("Unit") != clear_unit]
            st.session_state.excel_file = None
            st.rerun()

        if st.button("🗑️ Reset Semua Data", use_container_width=True):
            st.session_state.all_data = []
            st.session_state.excel_file = None
            st.rerun()

# Summary metrics
if st.session_state.all_data:
    summary_df = pd.DataFrame(st.session_state.all_data)
    total_rows = len(summary_df)
    loaded_units = summary_df["Unit"].nunique() if "Unit" in summary_df.columns else 0
    unmatched = int((summary_df.get("Supplier", pd.Series(dtype=str)).fillna("").astype(str).str.strip() == "").sum())
else:
    total_rows = loaded_units = unmatched = 0

m1, m2, m3 = st.columns(3)
m1.metric("Total baris", total_rows)
m2.metric("Unit terisi", f"{loaded_units}/{len(UNITS)}")
m3.metric("Supplier belum match", unmatched)

st.subheader("1. Upload Purchase Order")
unit = st.selectbox("Pilih unit", UNITS, key="selected_unit")

uploaded_files = st.file_uploader(
    f"Upload PDF untuk {unit}",
    type=["pdf"],
    accept_multiple_files=True,
    key=f"upload_{unit}",
    help="Bisa upload beberapa PDF sekaligus untuk unit yang sama.",
)

col_add, col_info = st.columns([1, 2])
with col_add:
    add_clicked = st.button(
        f"➕ Tambah Data {unit}",
        disabled=not uploaded_files,
        use_container_width=True,
        type="primary",
    )
with col_info:
    if uploaded_files:
        st.info(f"{len(uploaded_files)} PDF siap diproses untuk **{unit}**.")
    else:
        st.info("Pilih satu atau beberapa PDF purchase order, lalu klik **Tambah Data**.")

if add_clicked:
    added_rows = 0
    duplicate_rows = 0
    empty_files = []
    file_results = []
    existing_keys = {row_key(row) for row in st.session_state.all_data}

    with st.spinner(f"Membaca PDF untuk {unit}..."):
        for uploaded_file in uploaded_files:
            rows = parse_pdf(uploaded_file, unit, supplier_db)
            if not rows:
                empty_files.append(uploaded_file.name)
                file_results.append((uploaded_file.name, 0, 0))
                continue

            file_added = 0
            file_dupes = 0
            for row in rows:
                key = row_key(row)
                if key in existing_keys:
                    duplicate_rows += 1
                    file_dupes += 1
                    continue
                st.session_state.all_data.append(row)
                existing_keys.add(key)
                added_rows += 1
                file_added += 1
            file_results.append((uploaded_file.name, file_added, file_dupes))

    st.session_state.excel_file = None

    if added_rows:
        st.success(f"Berhasil menambahkan **{added_rows} baris** untuk {unit}.")
    if duplicate_rows:
        st.warning(f"{duplicate_rows} baris duplikat dilewati otomatis.")
    if empty_files:
        st.warning("Tidak ada item yang terbaca dari: " + ", ".join(empty_files))

    with st.expander("Lihat hasil per file"):
        result_df = pd.DataFrame(file_results, columns=["File", "Baris ditambah", "Duplikat dilewati"])
        st.dataframe(result_df, use_container_width=True, hide_index=True)

if st.session_state.all_data:
    st.divider()
    st.subheader("2. Review & Edit Data")
    df_preview = pd.DataFrame(st.session_state.all_data)

    f1, f2 = st.columns([1, 2])
    with f1:
        unit_filter = st.multiselect("Filter unit", UNITS, default=UNITS)
    with f2:
        search_text = st.text_input("Cari item / supplier / No PO", placeholder="contoh: chicken, PO/123, vendor...")

    mask = df_preview["Unit"].isin(unit_filter) if unit_filter else pd.Series(False, index=df_preview.index)
    if search_text.strip():
        needle = search_text.strip().lower()
        searchable = df_preview.astype(str).apply(lambda col: col.str.lower().str.contains(needle, na=False))
        mask &= searchable.any(axis=1)

    filtered_df = df_preview.loc[mask].copy()
    display_columns = [c for c in COLUMNS + ["PO Date"] if c in filtered_df.columns]
    st.caption(f"Menampilkan {len(filtered_df)} dari {len(df_preview)} baris.")
    st.dataframe(filtered_df[display_columns], use_container_width=True, hide_index=True, height=380)

    with st.expander("✏️ Edit seluruh data sebelum export"):
        st.caption("Perubahan di tabel ini akan dipakai saat membuat Excel.")
        edit_columns = [c for c in COLUMNS + ["PO Date"] if c in df_preview.columns]
        edited_df = st.data_editor(
            df_preview[edit_columns],
            use_container_width=True,
            hide_index=True,
            num_rows="dynamic",
            key="data_editor_all",
        )
        if st.button("Simpan Perubahan", use_container_width=True):
            st.session_state.all_data = edited_df.to_dict("records")
            st.session_state.excel_file = None
            st.success("Perubahan tersimpan.")
            st.rerun()

    unmatched_df = df_preview[df_preview["Supplier"].fillna("").astype(str).str.strip() == ""]
    if not unmatched_df.empty:
        with st.expander(f"⚠️ Supplier belum terdeteksi ({len(unmatched_df)} baris)"):
            st.dataframe(
                unmatched_df[[c for c in ["Unit", "Item", "No PO"] if c in unmatched_df.columns]],
                use_container_width=True,
                hide_index=True,
            )
            csv_unmatched = unmatched_df.to_csv(index=False).encode("utf-8-sig")
            st.download_button(
                "Download daftar supplier belum match (.csv)",
                data=csv_unmatched,
                file_name="supplier_belum_match.csv",
                mime="text/csv",
                use_container_width=True,
            )

    st.divider()
    st.subheader("3. Kirim PO ke Supplier via WhatsApp")
    st.caption("Data otomatis dikelompokkan per **unit + supplier**. Download gambar PO lalu buka chat supplier yang tepat.")

    wa_ready_df = pd.DataFrame(st.session_state.all_data)
    wa_ready_df["Supplier"] = wa_ready_df["Supplier"].fillna("").astype(str).str.strip()
    wa_ready_df = wa_ready_df[wa_ready_df["Supplier"] != ""]

    if wa_ready_df.empty:
        st.info("Belum ada supplier yang terdeteksi. Upload database supplier atau isi supplier di menu Review & Edit.")
    else:
        group_keys = ["Unit", "Supplier"]
        grouped = list(wa_ready_df.groupby(group_keys, sort=False))
        st.write(f"**{len(grouped)} supplier/unit siap dibuatkan PO WhatsApp.**")

        for (wa_unit, wa_supplier), wa_group in grouped:
            phone = lookup_supplier_phone(wa_supplier, supplier_db)
            po_image = build_supplier_po_image(wa_unit, wa_supplier, wa_group)
            safe_supplier = re.sub(r"[^A-Za-z0-9_-]+", "_", wa_supplier).strip("_") or "supplier"
            safe_unit = re.sub(r"[^A-Za-z0-9_-]+", "_", wa_unit).strip("_") or "unit"
            file_name = f"PO_{safe_unit}_{safe_supplier}.png"

            with st.expander(f"💬 {wa_supplier} · {wa_unit} · {len(wa_group)} item"):
                st.image(po_image.getvalue(), caption=f"Preview PO untuk {wa_supplier}", use_container_width=True)
                dcol, wcol = st.columns(2)
                with dcol:
                    st.download_button(
                        "🖼️ Download Gambar PO",
                        data=po_image.getvalue(),
                        file_name=file_name,
                        mime="image/png",
                        use_container_width=True,
                        key=f"download_wa_{safe_unit}_{safe_supplier}",
                    )
                with wcol:
                    if phone:
                        message = build_whatsapp_message(wa_unit, wa_supplier, wa_group)
                        wa_url = f"https://wa.me/{phone}?text={quote(message)}"
                        st.link_button(
                            "💚 Buka Chat WhatsApp",
                            wa_url,
                            use_container_width=True,
                        )
                    else:
                        st.button(
                            "💚 Nomor WA belum ada",
                            disabled=True,
                            use_container_width=True,
                            key=f"missing_wa_{safe_unit}_{safe_supplier}",
                        )
                        st.caption("Isi kolom **NO WA** pada DATABASE SUP.xlsx untuk supplier ini.")

                if phone:
                    st.caption(f"Nomor tujuan: +{phone}")
                st.info("WhatsApp akan membuka chat dan mengisi pesan otomatis. Gambar PO tetap perlu kamu attach dari hasil download karena wa.me tidak bisa menempelkan file otomatis.")

    st.divider()
    st.subheader("4. Convert & Download Excel")
    c1, c2 = st.columns([1, 1])
    with c1:
        if st.button("⚡ Convert Semua Unit", use_container_width=True, type="primary"):
            st.session_state.excel_file = build_excel(st.session_state.all_data)
            st.success("Excel berhasil dibuat. Setiap unit berada di sheet terpisah.")

    with c2:
        if st.session_state.excel_file:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M")
            st.download_button(
                "⬇️ Download Excel",
                data=st.session_state.excel_file,
                file_name=f"hasil_final_{timestamp}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        else:
            st.button("⬇️ Download Excel", disabled=True, use_container_width=True)
else:
    st.divider()
    st.info("Belum ada data. Mulai dari memilih unit dan upload PDF di atas.")
